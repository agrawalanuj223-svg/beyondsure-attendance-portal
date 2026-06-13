import mysql.connector
from datetime import datetime
import calendar
from flask import Flask, request, jsonify, render_template, session, redirect, url_for, send_file
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.secret_key = "beyondsure_super_secret" 

def get_db_connection():
    return mysql.connector.connect(
        host="localhost", user="root", password="anuj", 
        database="beyondsure_hr", use_pure=True, charset="utf8"
    )

# --- LOGIN & SESSION MGMT ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM Employees WHERE username=%s AND password=%s", (username, password))
        user = cursor.fetchone()
        db.close()

        if user:
            session['user_id'] = user['emp_id']
            session['role'] = user.get('role', 'Employee') 
            session['name'] = user['name']
            return redirect(url_for('home'))
            
        return render_template('login.html', error="Invalid credentials.")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# --- MAIN DASHBOARD ---
@app.route('/')
def home():
    if 'role' not in session: return redirect(url_for('login'))
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    
    if session['role'] == 'Admin':
        cursor.execute("SELECT emp_id, emp_code, name, role FROM Employees")
        all_employees = cursor.fetchall()
        cursor.execute("SELECT lr.request_id, e.name, lr.leave_type, lr.request_date FROM LeaveRequests lr JOIN Employees e ON lr.emp_id = e.emp_id WHERE lr.status = 'Pending'")
        pending_requests = cursor.fetchall()
        return render_template('index.html', employees=all_employees, pending_requests=pending_requests)
        
    elif session['role'] == 'Employee':
        cursor.execute("SELECT * FROM LeaveRequests WHERE emp_id = %s ORDER BY request_date DESC", (session['user_id'],))
        return render_template('employee.html', name=session['name'], requests=cursor.fetchall())

# --- EXCEL REPORT GENERATION ENGINE ---
@app.route('/download_report', methods=['POST'])
def download_report():
    if session.get('role') != 'Admin': return "Unauthorized", 403
    month = int(request.form['month'])
    year = int(request.form['year'])
    
    db = get_db_connection()
    query = """
        SELECT e.emp_id AS `EMP ID`, e.name AS `EMPLOYEE NAME`, a.record_date, a.status, a.in_time, a.out_time
        FROM Attendance a 
        JOIN Employees e ON a.emp_id = e.emp_id 
        WHERE MONTH(a.record_date) = %s AND YEAR(a.record_date) = %s
        ORDER BY a.record_date, e.name
    """
    df = pd.read_sql(query, db, params=(month, year))
    db.close()
    
    if df.empty:
        return "No records found for the chosen month and year.", 404

    # =========================================================
    # SHEET 1: THE MASTER MATRIX & PAYROLL MATH
    # =========================================================
    df['day'] = pd.to_datetime(df['record_date']).dt.day
    pivot_df = df.pivot(index='EMPLOYEE NAME', columns='day', values='status')
    
    _, days_in_month = calendar.monthrange(year, month)
    for day in range(1, days_in_month + 1):
        if day not in pivot_df.columns:
            pivot_df[day] = None
            
    day_cols = list(range(1, days_in_month + 1))
    calc_df = pivot_df[day_cols].map(lambda x: str(x).lower() if pd.notna(x) else "")
    
    pivot_df['WFH'] = (calc_df == 'wfh').sum(axis=1)
    pivot_df['CO used'] = (calc_df == 'co').sum(axis=1)
    pivot_df['AB'] = (calc_df == 'ab').sum(axis=1)
    pivot_df['Half Day'] = (calc_df == 'hd').sum(axis=1) * 0.5
    
    pivot_df['PL used'] = (calc_df == 'pl').sum(axis=1) + (calc_df == 'leave - pl').sum(axis=1)
    pivot_df['CL used'] = (calc_df == 'cl').sum(axis=1) + (calc_df == 'leave - cl').sum(axis=1)
    pivot_df['SL used'] = (calc_df == 'sl').sum(axis=1) + (calc_df == 'leave - sl').sum(axis=1)
    pivot_df['RH used'] = (calc_df == 'rh').sum(axis=1) + (calc_df == 'leave - rh').sum(axis=1)
    
    pivot_df['BL'] = (calc_df == 'bl').sum(axis=1) + (calc_df == 'leave - bl').sum(axis=1)
    pivot_df['AL'] = (calc_df == 'al').sum(axis=1) + (calc_df == 'leave - al').sum(axis=1)
    
    pivot_df['Weekly Off'] = (calc_df == 'wo').sum(axis=1)
    pivot_df['Holiday'] = (calc_df == 'h').sum(axis=1)

    pivot_df['Total Leaves'] = (pivot_df['PL used'] + pivot_df['CL used'] + pivot_df['SL used'] + 
                                pivot_df['CO used'] + pivot_df['RH used'] + pivot_df['AB'] + pivot_df['Half Day'])
    
    pivot_df['LOP'] = (pivot_df['Total Leaves'] - pivot_df['CO used'] - 1.5).clip(lower=0)
    pivot_df['Present Days'] = (calc_df == 'p').sum(axis=1) + pivot_df['WFH'] + pivot_df['Half Day']
    pivot_df['Payable Days'] = days_in_month - pivot_df['LOP']
    
    summary_cols = ['WFH', 'CO used', 'AB', 'Half Day', 'Total Leaves', 'PL used', 'RH used', 'BL', 'AL', 'LOP', 'Present Days', 'Payable Days', 'Weekly Off', 'Holiday']
    final_cols = day_cols + summary_cols
    pivot_df = pivot_df[final_cols].fillna("-")

    # =========================================================
    # SHEET 2: THE DETAILED TIME LOGS
    # =========================================================
    log_df = df[['EMP ID', 'EMPLOYEE NAME', 'record_date', 'status', 'in_time', 'out_time']].copy()
    log_df['record_date'] = pd.to_datetime(log_df['record_date']).dt.strftime('%Y-%m-%d')
    log_df['status'] = log_df['status'].str.upper()

    log_df['IN TIME'] = log_df['in_time'].apply(lambda x: str(x).split()[-1] if pd.notnull(x) else "-")
    log_df['OUT TIME'] = log_df['out_time'].apply(lambda x: str(x).split()[-1] if pd.notnull(x) else "-")

    log_df = log_df.drop(columns=['in_time', 'out_time'])
    log_df.rename(columns={'record_date': 'DATE', 'status': 'STATUS'}, inplace=True)

    # =========================================================
    # EXPORT WITH DYNAMIC HEADER BANNERS (MERGED & CENTERED)
    # =========================================================
    output = io.BytesIO()
    month_name = calendar.month_name[month]
    headline_text = f"Attendance Sheet for the month of {month_name} {year}"
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pivot_df.to_excel(writer, sheet_name='MASTER ATTENDANCE', startrow=3)
        log_df.to_excel(writer, sheet_name='DAILY TIME LOGS', index=False, startrow=3)
        
        workbook = writer.book
        header_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1A472A', end_color='1A472A', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Sheet 1 Branding
        ws1 = workbook['MASTER ATTENDANCE']
        ws1.merge_cells('A1:AL2')
        ws1['A1'] = headline_text
        ws1['A1'].font = header_font
        ws1['A1'].fill = header_fill
        ws1['A1'].alignment = center_alignment
        for row in ws1['A1:AL2']:
            for cell in row:
                cell.fill = header_fill

        # Sheet 2 Branding
        ws2 = workbook['DAILY TIME LOGS']
        ws2.merge_cells('A1:F2')
        ws2['A1'] = headline_text
        ws2['A1'].font = header_font
        ws2['A1'].fill = header_fill
        ws2['A1'].alignment = center_alignment
        for row in ws2['A1:F2']:
            for cell in row:
                cell.fill = header_fill

    output.seek(0)
    return send_file(output, download_name=f"BEYONDSURE_REPORT_{month}_{year}.xlsx", as_attachment=True)

# --- API ROUTES ---
@app.route('/api/mark_attendance', methods=['POST'])
def mark_attendance():
    data = request.json
    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute("INSERT INTO Attendance (emp_id, record_date, status, in_time, out_time) VALUES (%s, %s, %s, %s, %s) ON DUPLICATE KEY UPDATE status=VALUES(status), in_time=VALUES(in_time), out_time=VALUES(out_time)", 
                   (data['emp_id'], data['date'], data['status'], data.get('in_time') or None, data.get('out_time') or None))
    db.commit(); db.close()
    return jsonify({"success": True})

@app.route('/api/add_employee', methods=['POST'])
def add_employee():
    if session.get('role') != 'Admin': return "Unauthorized", 403
    db = get_db_connection()
    cursor = db.cursor()
    
    role_choice = request.form['role']
    emp_code = request.form.get('emp_code')
    name = request.form['name']
    username = request.form['username']
    password = request.form['password']

    try:
        cursor.execute(
            "INSERT INTO Employees (emp_code, name, username, password, role) VALUES (%s, %s, %s, %s, %s)", 
            (emp_code if emp_code and emp_code.strip() else None, name, username, password, role_choice)
        )
        db.commit()
    except mysql.connector.Error as err:
        print(f"Database Error while adding account profile: {err}")
    finally:
        db.close()
    return redirect(url_for('manage_employees'))

@app.route('/api/edit_employee', methods=['POST'])
def edit_employee():
    if session.get('role') != 'Admin': return "Unauthorized", 403
    db = get_db_connection()
    cursor = db.cursor()
    
    emp_id = request.form['emp_id']
    emp_code = request.form.get('emp_code')
    name = request.form['name']
    role = request.form['role']
    password = request.form['password']

    try:
        if password and password.strip():
            cursor.execute(
                "UPDATE Employees SET name=%s, role=%s, emp_code=%s, password=%s WHERE emp_id=%s",
                (name, role, emp_code if emp_code and emp_code.strip() else None, password, emp_id)
            )
        else:
            cursor.execute(
                "UPDATE Employees SET name=%s, role=%s, emp_code=%s WHERE emp_id=%s",
                (name, role, emp_code if emp_code and emp_code.strip() else None, emp_id)
            )
        db.commit()
    except mysql.connector.Error as err:
        print(f"Database Error during update execution: {err}")
    finally:
        db.close()
    return redirect(url_for('manage_employees'))

@app.route('/api/delete_employee/<int:emp_id>', methods=['POST'])
def delete_employee(emp_id):
    if session.get('role') != 'Admin': return "Unauthorized", 403
    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute("DELETE FROM Employees WHERE emp_id = %s", (emp_id,))
    db.commit(); db.close()
    return redirect(url_for('manage_employees'))

@app.route('/api/request_leave', methods=['POST'])
def request_leave():
    if 'role' not in session or session['role'] != 'Employee':
        return "Unauthorized", 403
        
    db = get_db_connection()
    cursor = db.cursor()
    emp_id = session['user_id']
    leave_type = request.form['leave_type']
    request_date = request.form['request_date']
    
    try:
        cursor.execute(
            "INSERT INTO LeaveRequests (emp_id, leave_type, request_date, status) VALUES (%s, %s, %s, 'Pending')",
            (emp_id, leave_type, request_date)
        )
        db.commit()
    except mysql.connector.Error as err:
        print(f"Database Error: {err}")
    finally:
        db.close()
    return redirect(url_for('home'))

# --- MANAGEMENT PAGES ---
@app.route('/manage_employees')
def manage_employees():
    if session.get('role') != 'Admin': return redirect(url_for('home'))
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM Employees")
    employees = cursor.fetchall()
    db.close()
    return render_template('manage_employees.html', employees=employees)

@app.route('/reports')
def reports():
    if session.get('role') != 'Admin': return redirect(url_for('home'))
    db = get_db_connection()
    query = """
        SELECT a.record_date, e.name, a.status 
        FROM Attendance a 
        JOIN Employees e ON a.emp_id = e.emp_id 
        ORDER BY a.record_date DESC LIMIT 100
    """
    df = pd.read_sql(query, db)
    db.close()
    if not df.empty:
        df['record_date'] = df['record_date'].astype(str)
    return render_template('reports.html', history=df.to_dict(orient='records'))

@app.route('/leave_approvals')
def leave_approvals():
    if session.get('role') != 'Admin': return redirect(url_for('home'))
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    
    query = """
        SELECT lr.request_id, e.name, e.emp_code, lr.leave_type, lr.request_date, lr.status 
        FROM LeaveRequests lr 
        JOIN Employees e ON lr.emp_id = e.emp_id 
        ORDER BY CASE WHEN lr.status = 'Pending' THEN 1 ELSE 2 END, lr.request_date DESC
    """
    cursor.execute(query)
    requests = cursor.fetchall()
    db.close()
    return render_template('leave_approvals.html', requests=requests)

@app.route('/api/resolve_leave', methods=['POST'])
def resolve_leave():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    action = request.form['action'] 
    req_id = request.form['request_id']
    
    cursor.execute("UPDATE LeaveRequests SET status=%s WHERE request_id=%s", (action, req_id))
    
    if action == 'Approved':
        cursor.execute("SELECT emp_id, leave_type, request_date FROM LeaveRequests WHERE request_id=%s", (req_id,))
        req = cursor.fetchone()
        db_status = f"Leave - {req['leave_type']}"
        cursor.execute("INSERT INTO Attendance (emp_id, record_date, status) VALUES (%s, %s, %s) ON DUPLICATE KEY UPDATE status = VALUES(status)", 
                       (req['emp_id'], req['request_date'], db_status))
    db.commit()
    db.close()
    return redirect(url_for('leave_approvals'))

@app.context_processor
def inject_pending_count():
    if session.get('role') == 'Admin' and request.endpoint != 'static':
        try:
            db = get_db_connection()
            cursor = db.cursor(buffered=True) 
            cursor.execute("SELECT COUNT(*) FROM LeaveRequests WHERE status = 'Pending'")
            row = cursor.fetchone()
            db.close()
            if row:
                return dict(pending_count=int(row[0]))
        except Exception as e:
            return dict(pending_count=0)
    return dict(pending_count=0)

if __name__ == '__main__':
    app.run(debug=True, use_reloader=False, port=5001)